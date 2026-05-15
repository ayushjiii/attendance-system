import csv
import calendar
from datetime import date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum
from attendance.models import AttendanceRecord
from leave_management.models import LeaveRequest
from accounts.models import Employee


@login_required
def monthly_report_view(request):
    if not request.user.is_admin_role:
        return redirect('attendance:dashboard')

    today = date.today()
    month_str = request.GET.get('month', today.strftime('%Y-%m'))

    try:
        year, month = month_str.split('-')
        year, month = int(year), int(month)
    except (ValueError, AttributeError):
        year, month = today.year, today.month

    # Count working days (Mon-sat)
    cal = calendar.monthcalendar(year, month)
    working_days = sum(1 for week in cal for day in week[:6] if day != 0)

    try:
        records = AttendanceRecord.objects.filter(
            date__year=year, date__month=month
        ).select_related('employee')
    except Exception:
        records = AttendanceRecord.objects.none()

    employees = Employee.objects.filter(role='employee')
    report_data = []

    for emp in employees:
        emp_records = records.filter(employee=emp)
        present = emp_records.filter(attendance_status='present').count()
        late = emp_records.filter(attendance_status='late').count()
        total_hours = emp_records.aggregate(total=Sum('total_working_hours'))['total'] or 0
        leaves_approved = LeaveRequest.objects.filter(
            employee=emp,
            leave_date__year=year,
            leave_date__month=month,
            status='approved'
        ).count()

        report_data.append({
            'employee': emp,
            'present_days': present,
            'late_days': late,
            'absent_days': working_days - present - leaves_approved,
            'approved_leaves': leaves_approved,
            'total_hours': round(total_hours, 2),
        })

    return render(request, 'reports/monthly_report.html', {
        'report_data': report_data,
        'month_str': month_str,
        'working_days': working_days,
    })

@login_required
def export_csv_view(request):
    if not request.user.is_admin_role:
        return redirect('attendance:dashboard')

    month_str = request.GET.get('month', date.today().strftime('%Y-%m'))
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_{month_str}.csv"'

    writer = csv.writer(response)
    writer.writerow(['Employee ID', 'Name', 'Department', 'Date',
                     'Check In', 'Check Out', 'Hours Worked', 'Status', 'IP Address'])

    try:
        year, month = month_str.split('-')
        records = AttendanceRecord.objects.filter(
            date__year=int(year), date__month=int(month)
        ).select_related('employee').order_by('employee__last_name', 'date')
    except (ValueError, AttributeError):
        records = AttendanceRecord.objects.none()

    for r in records:
        writer.writerow([
            r.employee.employee_id,
            r.employee.get_full_name(),
            r.employee.department,
            r.date,
            r.check_in_time or '',
            r.check_out_time or '',
            r.total_working_hours or '',
            r.get_attendance_status_display(),
            r.ip_address or '',
        ])

    return response

@login_required
def export_excel_view(request):
    if not request.user.is_admin_role:
        return redirect('attendance:dashboard')

    import openpyxl
    from openpyxl.styles import Font, PatternFill

    month_str = request.GET.get('month', date.today().strftime('%Y-%m'))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Attendance {month_str}'

    headers = ['Employee ID', 'Name', 'Department', 'Date',
               'Check In', 'Check Out', 'Hours', 'Status']
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    try:
        year, month = month_str.split('-')
        records = AttendanceRecord.objects.filter(
            date__year=int(year), date__month=int(month)
        ).select_related('employee').order_by('employee__last_name', 'date')
    except (ValueError, AttributeError):
        records = AttendanceRecord.objects.none()

    for r in records:
        ws.append([
            r.employee.employee_id,
            r.employee.get_full_name(),
            r.employee.department,
            str(r.date),
            str(r.check_in_time or ''),
            str(r.check_out_time or ''),
            float(r.total_working_hours or 0),
            r.get_attendance_status_display(),
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="attendance_{month_str}.xlsx"'
    wb.save(response)
    return response


  
@login_required
def employee_detail_report_view(request, employee_id):
    if not request.user.is_admin_role:
        return redirect('attendance:dashboard')

    from accounts.models import Employee
    import calendar

    employee = get_object_or_404(Employee, id=employee_id)
    today = date.today()
    month_str = request.GET.get('month', today.strftime('%Y-%m'))

    try:
        year, month = month_str.split('-')
        year, month = int(year), int(month)
    except (ValueError, AttributeError):
        year, month = today.year, today.month

    # All records for this employee this month
    records = AttendanceRecord.objects.filter(
        employee=employee,
        date__year=year,
        date__month=month
    ).order_by('date')

    # Build daily data for chart and table
    cal = calendar.monthcalendar(year, month)
    working_days = []
    for week in cal:
        for i, day in enumerate(week):
            if day != 0 and i < 6:  # Mon-sat
                working_days.append(date(year, month, day))

    # Check approved leaves
    approved_leaves = LeaveRequest.objects.filter(
        employee=employee,
        leave_date__year=year,
        leave_date__month=month,
        status='approved'
    ).values_list('leave_date', flat=True)

    daily_data = []
    chart_labels = []
    chart_hours = []
    chart_checkin = []

    for work_date in working_days:
        record = records.filter(date=work_date).first()

        if record and record.check_in_time:
            status = record.get_attendance_status_display()
            hours = float(record.total_working_hours or 0)
            checkin_minutes = record.check_in_time.hour * 60 + record.check_in_time.minute
            checkin_str = record.check_in_time.strftime('%I:%M %p')
            checkout_str = record.check_out_time.strftime('%I:%M %p') if record.check_out_time else '-'
            date_obj = record.check_in_time
        elif work_date in approved_leaves:
            status = 'On Leave'
            hours = 0
            checkin_minutes = None
            checkin_str = '-'
            checkout_str = '-'
            date_obj = None
        elif work_date > today:
            status = 'Upcoming'
            hours = 0
            checkin_minutes = None
            checkin_str = '-'
            checkout_str = '-'
        else:
            status = 'Absent'
            hours = 0
            checkin_minutes = None
            checkin_str = '-'
            checkout_str = '-'

        daily_data.append({
            'date': work_date,
            'day': work_date.strftime('%a'),
            'status': status,
            'check_in': checkin_str,
            'check_out': checkout_str,
            'hours': hours,
            'date_obj': date_obj,
        })

        chart_labels.append(work_date.strftime('%d %b'))
        chart_hours.append(hours)
        chart_checkin.append(checkin_minutes if checkin_minutes else 'null')

    # Summary stats
    present_count = sum(1 for d in daily_data if d['status'] not in ['Absent', 'On Leave', 'Upcoming'])
    absent_count = sum(1 for d in daily_data if d['status'] == 'Absent')
    leave_count = sum(1 for d in daily_data if d['status'] == 'On Leave')
    total_hours = sum(d['hours'] for d in daily_data)
    avg_hours = round(total_hours / present_count, 2) if present_count else 0

    # Average check-in time calculation (math-based instead of string-based)
    checkin_minute_list = [
        (d['date_obj'].hour * 60 + d['date_obj'].minute) 
        for d in daily_data if d.get('date_obj')
    ]
    
    if checkin_minute_list:
        avg_minutes = sum(checkin_minute_list) // len(checkin_minute_list)
        hour = avg_minutes // 60
        minute = avg_minutes % 60
        period = 'AM' if hour < 12 else 'PM'
        hour_12 = hour % 12 or 12
        avg_checkin = f"{hour_12:02d}:{minute:02d} {period}"
    else:
        avg_checkin = 'N/A'
        avg_minutes = None

    # Generate description
    if present_count == 0:
        description = f"{employee.get_full_name()} has no attendance records for this month."
    else:
        # Numeric punctuality check (threshold: 9:30 AM = 570 mins)
        if avg_minutes is not None:
            if avg_minutes <= 570:
                punctuality = "excellent"
            elif avg_minutes <= 600: # 10:00 AM
                punctuality = "good"
            else:
                punctuality = "needs improvement"
        else:
            punctuality = "N/A"

        description = (
            f"{employee.get_full_name()} was present for {present_count} out of "
            f"{len([d for d in daily_data if d['status'] != 'Upcoming'])} working days this month, "
            f"with {absent_count} absence(s) and {leave_count} approved leave(s). "
            f"Average working hours per day: {avg_hours} hrs. "
            f"Average check-in time: {avg_checkin} — punctuality is {punctuality}."
        )

    return render(request, 'reports/employee_detail.html', {
        'employee': employee,
        'daily_data': daily_data,
        'month_str': month_str,
        'present_count': present_count,
        'absent_count': absent_count,
        'leave_count': leave_count,
        'avg_hours': avg_hours,
        'avg_checkin': avg_checkin,
        'total_hours': total_hours,
        'description': description,
        'chart_labels': chart_labels,
        'chart_hours': chart_hours,
        'chart_checkin': chart_checkin,
    })